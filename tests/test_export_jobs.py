import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / 'scripts' / 'export_jobs.py'
spec = importlib.util.spec_from_file_location('export_jobs', MODULE_PATH)
export_jobs = importlib.util.module_from_spec(spec)
assert spec.loader is not None
spec.loader.exec_module(export_jobs)


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + '\n')


def test_build_export_rows_sorts_scores_and_enriches_listing_fields(tmp_path: Path):
    evaluations_dir = tmp_path / 'evaluations' / 'run-123'
    write_json(
        evaluations_dir / 'listing-b.json',
        {'listingId': 'listing-b', 'decision': 'drop', 'score': 40, 'reasoning': 'Too senior.'},
    )
    write_json(
        evaluations_dir / 'listing-a.json',
        {'listingId': 'listing-a', 'decision': 'keep', 'score': 91, 'reasoning': 'Strong fit.'},
    )
    write_json(
        evaluations_dir / 'listing-b.error.json',
        {'listingId': 'listing-b', 'error': 'This should be ignored.'},
    )

    listing_map = {
        'listing-a': {
            'id': 'listing-a',
            'title': 'Backend Engineer',
            'company': 'Acme',
            'location': 'Paris',
            'workMode': 'hybrid',
            'source': 'linkedin',
            'postedDate': '2026-03-28',
            'url': 'https://example.com/a',
            'summary': 'Java/Spring role',
        },
        'listing-b': {
            'id': 'listing-b',
            'title': 'Staff Engineer',
            'company': 'Globex',
            'location': 'Remote',
            'workMode': 'remote',
            'source': 'linkedin',
            'postedDate': '2026-03-27',
            'url': 'https://example.com/b',
            'summary': 'Very senior role',
        },
    }

    rows = export_jobs.build_export_rows(evaluations_dir, listing_map)

    assert [row['listingId'] for row in rows] == ['listing-a', 'listing-b']
    assert rows[0]['title'] == 'Backend Engineer'
    assert rows[0]['reasoning'] == 'Strong fit.'
    assert rows[1]['company'] == 'Globex'


def test_write_xlsx_outputs_expected_headers_and_values(tmp_path: Path):
    rows = [
        {
            'score': 88,
            'decision': 'keep',
            'title': 'Backend Engineer',
            'company': 'Acme',
            'location': 'Paris',
            'workMode': 'hybrid',
            'source': 'linkedin',
            'postedDate': '2026-03-28',
            'url': 'https://example.com/a',
            'reasoning': 'Strong fit.',
            'summary': 'Java/Spring role',
            'listingId': 'listing-a',
            'evaluationPath': '/tmp/listing-a.json',
        }
    ]
    xlsx_path = tmp_path / 'export.xlsx'

    export_jobs.write_xlsx(rows, xlsx_path, sheet_title='Evaluations')

    workbook = load_workbook(xlsx_path)
    sheet = workbook['Evaluations']
    assert sheet.cell(row=1, column=1).value == 'score'
    assert sheet.cell(row=2, column=1).value == '88'
    assert sheet.cell(row=2, column=3).value == 'Backend Engineer'
    assert sheet.cell(row=2, column=13).value == '/tmp/listing-a.json'
